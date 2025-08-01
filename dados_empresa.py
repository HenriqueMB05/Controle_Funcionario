from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import findpath
import os

wb = Workbook()
ws = wb.active

def addLine(caminho):
    with open(findpath.find(caminho), "r", encoding="utf-8") as file:
        for line in file:
            cell = line.strip().split(";")
            new_cell = [item.split() for item in cell]
            ws.append(new_cell)

def stl(name):
    header = [cell.value for cell in ws[1]]
    column_index = header.index(name) +1
    letter_column = get_column_letter(column_index)

    for row in range(2, ws.max_row+1):
        cell = ws.cell(row=row, column=column_index)
        cell.number_format = 'R$ #,##0.00'

ws.title = "Contatos"
ws.append(['Nome', 'Telefone', 'E-mail'])
addLine("contatos.txt")
    

ws = wb.create_sheet(title="Produtos")
ws.append(['Id','Produto', 'Preço'])
addLine("produtos.txt")
stl('Preço')

ws = wb.create_sheet(title="Vendas")
ws.append(['Data', 'Produto', 'Quantidade', 'Total'])
addLine("vendas.txt")
stl('Total')

wb.save("dados_empresa.xlsx")


os.system(f"start dados_empresa.xlsx")