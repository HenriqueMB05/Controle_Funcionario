from openpyxl import Workbook, load_workbook
from ReadPDF import *
import os

data = ["Nome", "Nascimento", "Sexo","CEP", "Bairro","Tipo lougradouro", "Lougradouro", "Número", "Complemento", "Cidade", "UF", "CPF", "RG", "Data de Emissão", "Orgão emissor", "UF"]
employer_data = []


# O nome por enquanto será pré-definido depois o usúario será responsavel por passar um arquivo já existente ou criar um novo
path_file = "Controle_Funcionario.xlsx"

try:
    wb = load_workbook(path_file)
    for i, row in enumerate(wb.active.iter_rows(values_only=True), start=1):
        if any(cell is not None for cell in row):
            line = i+1
except FileNotFoundError:
    wb = Workbook()    
    line = 2
    
ws = wb.active

for i, k in enumerate(data):
    ws.cell(row=1, column=i+1, value=data[i])

linha = find_line("Empregado ")

data = find_line("Data de nascimento").split(" ")
nascimento = data[0]


# Area do preenchimento da planilha 

wb.save(path_file)
os.system(f"start {path_file}")